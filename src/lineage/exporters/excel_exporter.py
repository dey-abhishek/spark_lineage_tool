"""Excel exporter for lineage data with multiple formatted sheets."""

from pathlib import Path
from typing import Dict, List
import json
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from lineage.lineage import LineageGraph, NodeType


class ExcelExporter:
    """Export lineage data to Excel with multiple formatted sheets."""
    
    def __init__(self, graph: LineageGraph, metrics: Dict = None):
        self.graph = graph
        self.metrics = metrics or {}
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    def _should_exclude_dataset(self, node) -> bool:
        """Check if dataset should be excluded from reports (only truly invalid entries)."""
        name = node.name.lower()
        urn = node.urn.lower()
        
        # Skip wildcard patterns
        if '*' in name or '*' in urn:
            return True
        
        # Skip temporary/staging paths
        temp_patterns = ['_tmp', '_temp', '_staging', '_checkpoint', '_success', '_logs', '_metadata']
        if any(pattern in name or pattern in urn for pattern in temp_patterns):
            return True
        
        # Skip very generic names
        generic_names = ['data', 'tmp', 'temp', 'staging', '-update']
        if name in generic_names:
            return True
        
        # Skip very short names (likely partial extractions)
        if len(name) < 3:
            return True
        
        # KEEP unresolved variables like hdfsDir, inputPath, etc.
        # Show them as "NA" or unresolved rather than hiding them
        # This provides visibility that a dataset exists but couldn't be fully resolved
        
        return False
    
    def _extract_table_from_jdbc_query(self, query: str) -> str:
        """Extract table name from JDBC SQL query."""
        if not query:
            return ""
        
        # Remove outer parentheses if present
        query = query.strip()
        while query.startswith('(') and query.endswith(')'):
            query = query[1:-1].strip()
        
        # Pattern 1: Simple table name at start (before WHERE/JOIN/ORDER)
        # Example: "CUSTOMERS WHERE ..." -> "CUSTOMERS"
        simple_pattern = r'^([a-zA-Z_][\w.]*?)(?:\s+WHERE|\s+JOIN|\s+ORDER|\s+GROUP|\s+LIMIT|\s*$)'
        match = re.search(simple_pattern, query, re.IGNORECASE)
        if match:
            table_name = match.group(1).strip()
            # Make sure it's not a SQL keyword
            if table_name.upper() not in ['SELECT', 'WITH', 'INSERT', 'UPDATE', 'DELETE']:
                return table_name.lower()
        
        # Pattern 2: SELECT ... FROM table_name
        from_pattern = r'FROM\s+([a-zA-Z_][\w.]*)'
        match = re.search(from_pattern, query, re.IGNORECASE)
        if match:
            table_name = match.group(1).strip()
            # Remove any trailing SQL keywords
            table_name = re.split(r'\s+WHERE|\s+JOIN|\s+ORDER', table_name, flags=re.IGNORECASE)[0]
            # Handle schema.table
            table_parts = table_name.split('.')
            if len(table_parts) > 1:
                # Return schema.table
                return '.'.join(table_parts[-2:]).lower()
            return table_name.lower()
        
        # Pattern 3: INTO clause (for INSERT)
        into_pattern = r'INTO\s+([a-zA-Z_][\w.]*)'
        match = re.search(into_pattern, query, re.IGNORECASE)
        if match:
            table_name = match.group(1).strip()
            table_parts = table_name.split('.')
            if len(table_parts) > 1:
                return '.'.join(table_parts[-2:]).lower()
            return table_name.lower()
        
        # If no pattern found, return truncated query
        return f"(query:{query[:30]}...)"
    
    def export(self, output_path: Path) -> None:
        """Export to Excel workbook with multiple sheets."""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create sheets
        self._create_summary_sheet(wb)
        self._create_priority_sheet(wb)
        self._create_source_analysis_sheet(wb)
        self._create_pivot_ready_sheet(wb)
        self._create_datasets_sheet(wb)
        self._create_jobs_sheet(wb)
        self._create_lineage_sheet(wb)
        self._create_metrics_sheet(wb)
        
        # Save workbook
        output_file = output_path / "lineage_report.xlsx"
        wb.save(output_file)
        print(f"Exported Excel report to {output_file}")
    
    def _create_summary_sheet(self, wb: Workbook) -> None:
        """Create summary overview sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Lineage Analysis Summary Report"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # Statistics
        row = 3
        graph_stats = self.graph.get_stats()
        stats = [
            ("Total Nodes", graph_stats.get('total_nodes', 0)),
            ("Total Edges", graph_stats.get('total_edges', 0)),
            ("Dataset Nodes", len(self.graph.get_nodes_by_type(NodeType.DATASET))),
            ("Job Nodes", len(self.graph.get_nodes_by_type(NodeType.JOB))),
            ("Module Nodes", len(self.graph.get_nodes_by_type(NodeType.MODULE))),
        ]
        
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Count"
        self._style_header_row(ws, row, 2)
        
        for stat_name, stat_value in stats:
            row += 1
            ws[f'A{row}'] = stat_name
            ws[f'B{row}'] = stat_value
        
        # Metrics summary if available
        if self.metrics:
            row += 2
            ws[f'A{row}'] = "Priority Analysis"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            wave_counts = {}
            for metric in self.metrics.values():
                wave = metric.migration_wave
                wave_counts[wave] = wave_counts.get(wave, 0) + 1
            
            ws[f'A{row}'] = "Migration Wave"
            ws[f'B{row}'] = "Datasets"
            self._style_header_row(ws, row, 2)
            
            for wave in sorted(wave_counts.keys()):
                row += 1
                ws[f'A{row}'] = f"Wave {wave}"
                ws[f'B{row}'] = wave_counts[wave]
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _create_priority_sheet(self, wb: Workbook) -> None:
        """Create top priority datasets sheet."""
        ws = wb.create_sheet("Top Priorities")
        
        # Header
        ws['A1'] = "Top Priority Datasets for Migration"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')
        
        # Column headers
        headers = ["Rank", "Dataset Name", "Priority Score", "Wave", "Downstream Reach", "Fan Out", "Confidence"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Get top datasets
        if self.metrics:
            sorted_metrics = sorted(
                self.metrics.items(),
                key=lambda x: x[1].priority_score,
                reverse=True
            )[:50]  # Top 50
            
            row = 4
            for rank, (node_id, metric) in enumerate(sorted_metrics, 1):
                # Find node by ID
                node = self.graph.get_node_by_id(node_id)
                
                if not node:
                    continue
                
                # Determine wave color
                if metric.migration_wave == 1:
                    fill_color = "FFE6E6"  # Light red
                elif metric.migration_wave == 2:
                    fill_color = "FFF4E6"  # Light yellow
                else:
                    fill_color = "E6F4EA"  # Light green
                
                ws.cell(row, 1, rank)
                ws.cell(row, 2, node.name)
                ws.cell(row, 3, round(metric.priority_score, 2))
                ws.cell(row, 4, f"Wave {metric.migration_wave}")
                ws.cell(row, 5, metric.downstream_reach)
                ws.cell(row, 6, metric.fan_out)
                ws.cell(row, 7, round(metric.avg_confidence, 2))
                
                # Color-code by wave
                for col in range(1, 8):
                    ws.cell(row, col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                row += 1
        
        # Auto-size columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_source_analysis_sheet(self, wb: Workbook) -> None:
        """Create source analysis sheet showing relationships per source."""
        ws = wb.create_sheet("Source Analysis")
        
        # Header
        ws['A1'] = "Lineage Relationships by Source"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:H1')
        
        # Column headers
        headers = ["Source File/System", "Type", "Datasets Read", "Datasets Written", 
                   "Total Relationships", "Avg Confidence", "Languages", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Aggregate data by source
        source_stats = {}
        
        # Analyze jobs
        for node in self.graph.get_nodes_by_type(NodeType.JOB):
            source_file = node.metadata.get('source_file', 'Unknown')
            
            # Determine type from file extension
            if source_file.endswith('.py'):
                source_type = 'PySpark'
            elif source_file.endswith('.scala'):
                source_type = 'Scala Spark'
            elif source_file.endswith('.hql') or source_file.endswith('.sql'):
                source_type = 'Hive SQL'
            elif source_file.endswith('.sh'):
                source_type = 'Shell Script'
            elif source_file.endswith('.json'):
                source_type = 'NiFi Flow'
            else:
                source_type = 'Other'
            
            if source_file not in source_stats:
                source_stats[source_file] = {
                    'type': source_type,
                    'reads': set(),
                    'writes': set(),
                    'confidences': [],
                    'languages': set([source_type])
                }
            
            # Count reads and writes
            for edge in self.graph.edges:
                if edge.source_node_id == node.node_id:
                    # This job writes to something
                    target = self.graph.get_node_by_id(edge.target_node_id)
                    if target and target.node_type == NodeType.DATASET:
                        source_stats[source_file]['writes'].add(target.name)
                        source_stats[source_file]['confidences'].append(edge.confidence)
                
                if edge.target_node_id == node.node_id:
                    # This job reads from something
                    source = self.graph.get_node_by_id(edge.source_node_id)
                    if source and source.node_type == NodeType.DATASET:
                        source_stats[source_file]['reads'].add(source.name)
                        source_stats[source_file]['confidences'].append(edge.confidence)
        
        # Sort by total relationships
        sorted_sources = sorted(
            source_stats.items(),
            key=lambda x: len(x[1]['reads']) + len(x[1]['writes']),
            reverse=True
        )
        
        # Populate data
        row = 4
        for source_file, stats in sorted_sources:
            total_rels = len(stats['reads']) + len(stats['writes'])
            avg_conf = sum(stats['confidences']) / len(stats['confidences']) if stats['confidences'] else 0
            
            # Determine status
            if avg_conf >= 0.8:
                status = "✓ High Confidence"
                status_color = "C6EFCE"  # Green
            elif avg_conf >= 0.6:
                status = "⚠ Medium Confidence"
                status_color = "FFEB9C"  # Yellow
            else:
                status = "✗ Low Confidence"
                status_color = "FFC7CE"  # Red
            
            # Short name for display
            short_name = source_file.split('/')[-1] if '/' in source_file else source_file
            
            ws.cell(row, 1, short_name)
            ws.cell(row, 1).hyperlink = f"#'All Jobs'!A1"  # Link to jobs sheet
            ws.cell(row, 1).font = Font(color="0563C1", underline="single")
            ws.cell(row, 2, stats['type'])
            ws.cell(row, 3, len(stats['reads']))
            ws.cell(row, 4, len(stats['writes']))
            ws.cell(row, 5, total_rels)
            ws.cell(row, 6, round(avg_conf, 2))
            ws.cell(row, 7, ', '.join(sorted(stats['languages'])))
            ws.cell(row, 8, status)
            
            # Color-code status
            ws.cell(row, 8).fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
            
            # Add tooltip with full path
            ws.cell(row, 1).comment = Comment(f"Full path: {source_file}", "Lineage Tool")
            
            row += 1
        
        # Add summary statistics at the bottom
        row += 2
        ws.cell(row, 1, "Summary Statistics").font = Font(bold=True, size=12)
        row += 1
        
        total_sources = len(source_stats)
        total_pyspark = sum(1 for s in source_stats.values() if 'PySpark' in s['type'])
        total_scala = sum(1 for s in source_stats.values() if 'Scala' in s['type'])
        total_hive = sum(1 for s in source_stats.values() if 'Hive' in s['type'])
        total_shell = sum(1 for s in source_stats.values() if 'Shell' in s['type'])
        
        summary_data = [
            ("Total Source Files", total_sources),
            ("PySpark Jobs", total_pyspark),
            ("Scala Spark Jobs", total_scala),
            ("Hive SQL Scripts", total_hive),
            ("Shell Scripts", total_shell),
        ]
        
        ws.cell(row, 1, "Category").font = Font(bold=True)
        ws.cell(row, 2, "Count").font = Font(bold=True)
        for cat, count in summary_data:
            row += 1
            ws.cell(row, 1, cat)
            ws.cell(row, 2, count)
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
    
    def _create_pivot_ready_sheet(self, wb: Workbook) -> None:
        """Create a flattened sheet perfect for pivot tables and analytics."""
        ws = wb.create_sheet("Analytics (Pivot-Ready)")
        
        # Header with explanation
        ws['A1'] = "Lineage Relationships - Source Dataset → Job → Target Dataset"
        ws['A1'].font = Font(size=12, bold=True, color="0563C1")
        ws.merge_cells('A1:K1')
        
        # Column headers - NEW structure with Source and Target
        headers = [
            "Source Dataset", "Source Type", "Source Schema.Table",
            "Job Name",
            "Relationship", 
            "Target Dataset", "Target Type", "Target Schema.Table",
            "Dataset URN/Path",
            "Fully Resolved", "Wave"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col, header)
            self._style_header_row(ws, 3, len(headers))
        
        # Add helpful column descriptions in row 2
        descriptions = [
            "Input dataset (for READ)",
            "Type of source",
            "database.table",
            "Job that reads/writes",
            "READ/WRITE/TRANSFORM",
            "Output dataset",
            "Type of target",
            "database.table",
            "Paths for files",
            "All variables resolved?",
            "Migration wave"
        ]
        for col, desc in enumerate(descriptions, 1):
            cell = ws.cell(2, col, desc)
            cell.font = Font(size=9, italic=True, color="666666")
            cell.alignment = Alignment(horizontal="center")
        
        # Flatten all relationships
        row = 4
        for edge in self.graph.edges:
            # Get source and target nodes
            source_node = self.graph.get_node_by_id(edge.source_node_id)
            target_node = self.graph.get_node_by_id(edge.target_node_id)
            
            if not source_node or not target_node:
                continue
            
            # Skip edges involving datasets that look like variable names
            if source_node.node_type == NodeType.DATASET and self._should_exclude_dataset(source_node):
                continue
            if target_node.node_type == NodeType.DATASET and self._should_exclude_dataset(target_node):
                continue
            
            # Determine relationship type and extract source/target/job
            source_dataset_node = None
            target_dataset_node = None
            job_node = None
            job_name = ""
            relationship = ""
            
            if source_node.node_type == NodeType.DATASET and target_node.node_type == NodeType.DATASET:
                # TRANSFORM: Source Dataset → Target Dataset (no job)
                source_dataset_node = source_node
                target_dataset_node = target_node
                relationship = "TRANSFORM"
                job_name = "(direct transformation)"
            elif source_node.node_type == NodeType.JOB and target_node.node_type == NodeType.DATASET:
                # WRITE: Job → Target Dataset
                job_node = source_node
                target_dataset_node = target_node
                relationship = "WRITE"
                source_file = job_node.metadata.get('source_file', 'Unknown')
                job_name = source_file.split('/')[-1] if '/' in source_file else source_file
            elif source_node.node_type == NodeType.DATASET and target_node.node_type == NodeType.JOB:
                # READ: Source Dataset → Job
                source_dataset_node = source_node
                job_node = target_node
                relationship = "READ"
                source_file = job_node.metadata.get('source_file', 'Unknown')
                job_name = source_file.split('/')[-1] if '/' in source_file else source_file
            else:
                continue
            
            # Extract source dataset details (for READ and TRANSFORM)
            source_dataset_name = ""
            source_dataset_type = ""
            source_schema_table = ""
            if source_dataset_node:
                source_dataset_name = source_dataset_node.name
                source_dataset_type = source_dataset_node.metadata.get('dataset_type', 'unknown')
                # Extract schema.table for source if it's a table
                if source_dataset_type in ['hive', 'table', 'jdbc', 'oracle', 'mysql', 'postgres', 'mssql']:
                    if '.' in source_dataset_name and not source_dataset_name.startswith('/'):
                        source_schema_table = source_dataset_name
            
            # Extract target dataset details (for WRITE and TRANSFORM)
            target_dataset_name = ""
            target_dataset_type = ""
            target_schema_table = ""
            target_dataset_urn = ""
            display_urn = ""
            if target_dataset_node:
                target_dataset_name = target_dataset_node.name
                target_dataset_urn = target_dataset_node.urn or target_dataset_name
                target_dataset_type = target_dataset_node.metadata.get('dataset_type', 'unknown')
                
                # For JDBC: clean up SQL queries
                if target_dataset_type in ['jdbc', 'oracle', 'mysql', 'postgres', 'mssql']:
                    if any(keyword in target_dataset_name.upper() for keyword in ['SELECT', 'FROM', 'WHERE', 'JOIN']) or len(target_dataset_name) > 50:
                        table_name = self._extract_table_from_jdbc_query(target_dataset_name)
                        if table_name and not table_name.startswith('(query:'):
                            target_dataset_name = table_name
                
                # Extract schema.table for target if it's a table
                if target_dataset_type in ['hive', 'table']:
                    if '.' in target_dataset_name and not target_dataset_name.startswith('/'):
                        target_schema_table = target_dataset_name
                    elif '.' in target_dataset_urn and not target_dataset_urn.startswith('/') and '://' not in target_dataset_urn:
                        target_schema_table = target_dataset_urn
                
                # For JDBC tables
                if target_dataset_type in ['jdbc', 'oracle', 'mysql', 'postgres', 'mssql']:
                    if '.' in target_dataset_name and not target_dataset_name.startswith('/'):
                        target_schema_table = target_dataset_name
                
                # Dataset URN/Path - ONLY for files
                original_urn = target_dataset_node.metadata.get('original_urn', '')
                resolved_urn = target_dataset_node.metadata.get('resolved_urn', target_dataset_urn)
                
                if target_dataset_type in ['hdfs', 'file', 'local', 'sftp', 's3', 'gcs']:
                    if original_urn and resolved_urn and original_urn != resolved_urn:
                        if '${' in original_urn and (':-' in original_urn or '$(date' in original_urn):
                            display_urn = f"{original_urn} → {resolved_urn}"
                        else:
                            display_urn = resolved_urn
                    else:
                        if resolved_urn and (resolved_urn.startswith('/') or '://' in resolved_urn):
                            display_urn = resolved_urn
                        elif target_dataset_urn and (target_dataset_urn.startswith('/') or '://' in target_dataset_urn):
                            display_urn = target_dataset_urn
                        elif target_dataset_name.startswith('/') or '://' in target_dataset_name:
                            display_urn = target_dataset_name
            
            # Determine if fully resolved
            fully_resolved = True
            if target_dataset_node:
                fully_resolved = target_dataset_node.metadata.get('fully_resolved', False)
                if not fully_resolved:
                    check_string = f"{target_dataset_name}{target_dataset_urn}{display_urn}"
                    if '${' in check_string or '$(' in check_string:
                        fully_resolved = False
                    else:
                        fully_resolved = True
            
            # Get metrics for wave
            wave = 0
            if target_dataset_node:
                metrics = self.metrics.get(target_dataset_node.node_id, None)
                wave = metrics.migration_wave if metrics else 0
            
            # Populate row - NEW structure (11 columns)
            ws.cell(row, 1, source_dataset_name[:100] if source_dataset_name else "")  # Source Dataset
            ws.cell(row, 2, source_dataset_type if source_dataset_name else "")          # Source Type
            ws.cell(row, 3, source_schema_table if source_schema_table else "")          # Source Schema.Table
            ws.cell(row, 4, job_name[:100])                                              # Job Name
            ws.cell(row, 5, relationship)                                                # Relationship
            ws.cell(row, 6, target_dataset_name[:100] if target_dataset_name else "")   # Target Dataset
            ws.cell(row, 7, target_dataset_type if target_dataset_name else "")          # Target Type
            ws.cell(row, 8, target_schema_table if target_schema_table else "")          # Target Schema.Table
            ws.cell(row, 9, display_urn[:200] if display_urn else "")                    # Dataset URN/Path
            ws.cell(row, 10, "Yes" if fully_resolved else "No")                         # Fully Resolved
            ws.cell(row, 11, f"Wave {wave}" if wave > 0 else "N/A")                     # Wave
            
            # Color-code fully resolved status
            if fully_resolved:
                ws.cell(row, 10).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws.cell(row, 10).fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
            
            row += 1
        
        # Add instructions
        row += 2
        ws.cell(row, 1, "💡 Column Explanations:").font = Font(bold=True, size=11, color="0563C1")
        row += 1
        explanations = [
            "• Source Dataset (Col 1): Input dataset that is READ by the job (empty for WRITE relationships)",
            "• Job Name (Col 4): The script/application that processes the data",
            "• Target Dataset (Col 6): Output dataset that is WRITTEN by the job (empty for READ relationships)",
            "• Relationship (Col 5): READ = job reads from source | WRITE = job writes to target | TRANSFORM = source transforms to target",
            "• Dataset URN/Path (Col 9): Full path for FILES ONLY (HDFS/local/SFTP/S3). Empty for tables.",
            "• Schema.Table (Col 3 & 8): For TABLES - database.table format (e.g., analytics.customers, prod.users)",
        ]
        for explanation in explanations:
            ws.cell(row, 1, explanation)
            ws.merge_cells(f'A{row}:H{row}')
            row += 1
        
        row += 1
        ws.cell(row, 1, "📊 How to Use for Analytics:").font = Font(bold=True, size=11, color="0563C1")
        row += 1
        instructions = [
            "1. Create Pivot Table: Select all data → Insert → Pivot Table",
            "2. Find job inputs: Filter Relationship=READ, then check Source Dataset column",
            "3. Find job outputs: Filter Relationship=WRITE, then check Target Dataset column",
            "4. Find transformations: Filter Relationship=TRANSFORM to see source→target dataset flows",
            "5. Analyze by job: Group by Job Name to see all inputs and outputs per job",
            "6. Hive table analysis: Filter Source/Target Type=hive, use Schema.Table columns",
            "7. File analysis: Filter Source/Target Type=hdfs/sftp/s3, use Dataset URN/Path column",
        ]
        for instruction in instructions:
            ws.cell(row, 1, instruction)
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
        
        # Auto-size columns (11 columns now)
        ws.column_dimensions['A'].width = 35  # Source Dataset
        ws.column_dimensions['B'].width = 12  # Source Type
        ws.column_dimensions['C'].width = 30  # Source Schema.Table
        ws.column_dimensions['D'].width = 35  # Job Name
        ws.column_dimensions['E'].width = 12  # Relationship
        ws.column_dimensions['F'].width = 35  # Target Dataset
        ws.column_dimensions['G'].width = 12  # Target Type
        ws.column_dimensions['H'].width = 30  # Target Schema.Table
        ws.column_dimensions['I'].width = 55  # Dataset URN/Path
        ws.column_dimensions['J'].width = 15  # Fully Resolved
        ws.column_dimensions['K'].width = 10  # Wave
    
    def _create_datasets_sheet(self, wb: Workbook) -> None:
        """Create all datasets sheet."""
        ws = wb.create_sheet("All Datasets")
        
        # Headers
        headers = ["Dataset Name", "URN", "Type", "Fully Resolved", "Fact Count", "Upstream Jobs", "Downstream Jobs"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            self._style_header_row(ws, 1, len(headers))
        
        # Data - filter out variable-like names
        row = 2
        for node in self.graph.get_nodes_by_type(NodeType.DATASET):
            # Skip datasets that look like unresolved variable names
            if self._should_exclude_dataset(node):
                continue
            
            ws.cell(row, 1, node.name)
            ws.cell(row, 2, node.urn)
            ws.cell(row, 3, node.metadata.get('dataset_type', 'unknown'))
            ws.cell(row, 4, "Yes" if node.metadata.get('fully_resolved', False) else "No")
            ws.cell(row, 5, node.metadata.get('fact_count', 0))
            ws.cell(row, 6, self.graph.get_fan_in(node.node_id))
            ws.cell(row, 7, self.graph.get_fan_out(node.node_id))
            row += 1
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    def _create_jobs_sheet(self, wb: Workbook) -> None:
        """Create all jobs sheet."""
        ws = wb.create_sheet("All Jobs")
        
        # Headers
        headers = ["Job Name", "Source File", "Reads From", "Writes To"]
        for col, header in enumerate(headers, 1):
            self._style_header_row(ws, 1, len(headers))
            ws.cell(1, col, header)
        
        # Data
        row = 2
        for node in self.graph.get_nodes_by_type(NodeType.JOB):
            source_file = node.metadata.get('source_file', '')
            
            # Count inputs/outputs
            upstream_count = len([e for e in self.graph.edges if e.target_node_id == node.node_id])
            downstream_count = len([e for e in self.graph.edges if e.source_node_id == node.node_id])
            
            ws.cell(row, 1, node.name)
            ws.cell(row, 2, source_file)
            ws.cell(row, 3, f"{upstream_count} datasets")
            ws.cell(row, 4, f"{downstream_count} datasets")
            row += 1
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 30
    
    def _create_lineage_sheet(self, wb: Workbook) -> None:
        """Create lineage relationships sheet."""
        ws = wb.create_sheet("Lineage Edges")
        
        # Headers
        headers = ["Source", "Relationship", "Target", "Confidence", "Evidence"]
        for col, header in enumerate(headers, 1):
            self._style_header_row(ws, 1, len(headers))
            ws.cell(1, col, header)
        
        # Data (limit to top edges by confidence)
        sorted_edges = sorted(self.graph.edges, key=lambda e: e.confidence, reverse=True)[:1000]
        
        row = 2
        for edge in sorted_edges:
            source_node = self.graph.get_node_by_id(edge.source_node_id)
            target_node = self.graph.get_node_by_id(edge.target_node_id)
            
            if not source_node or not target_node:
                continue
            
            ws.cell(row, 1, source_node.name)
            ws.cell(row, 2, edge.edge_type.value)
            ws.cell(row, 3, target_node.name)
            ws.cell(row, 4, round(edge.confidence, 2))
            ws.cell(row, 5, edge.evidence[:100] if edge.evidence else "")
            
            # Color-code by confidence
            if edge.confidence >= 0.8:
                fill_color = "C6EFCE"  # Green
            elif edge.confidence >= 0.6:
                fill_color = "FFEB9C"  # Yellow
            else:
                fill_color = "FFC7CE"  # Red
            
            ws.cell(row, 4).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            row += 1
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
    
    def _create_metrics_sheet(self, wb: Workbook) -> None:
        """Create detailed metrics sheet."""
        if not self.metrics:
            return
        
        ws = wb.create_sheet("Detailed Metrics")
        
        # Headers
        headers = ["Dataset", "Fan In", "Fan Out", "Downstream Reach", "Write Jobs", 
                   "Avg Confidence", "Priority Score", "Migration Wave"]
        for col, header in enumerate(headers, 1):
            self._style_header_row(ws, 1, len(headers))
            ws.cell(1, col, header)
        
        # Data
        sorted_metrics = sorted(
            self.metrics.items(),
            key=lambda x: x[1].priority_score,
            reverse=True
        )
        
        row = 2
        for node_id, metric in sorted_metrics:
            node = self.graph.get_node_by_id(node_id)
            if not node:
                continue
            
            ws.cell(row, 1, node.name)
            ws.cell(row, 2, metric.fan_in)
            ws.cell(row, 3, metric.fan_out)
            ws.cell(row, 4, metric.downstream_reach)
            ws.cell(row, 5, metric.write_job_count)
            ws.cell(row, 6, round(metric.avg_confidence, 2))
            ws.cell(row, 7, round(metric.priority_score, 2))
            ws.cell(row, 8, f"Wave {metric.migration_wave}")
            row += 1
        
        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _style_header_row(self, ws, row: int, num_cols: int) -> None:
        """Apply consistent styling to header rows."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

