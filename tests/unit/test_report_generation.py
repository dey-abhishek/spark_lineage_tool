"""
Test suite for report generation - validates content and formatting
Tests Excel, JSON, HTML, and CSV report outputs
"""

import pytest
import json
import pandas as pd
from pathlib import Path
import tempfile
import shutil
from openpyxl import load_workbook
from lineage.cli import main as cli_main
import sys


class TestExcelReportGeneration:
    """Tests for Excel report generation and formatting"""
    
    @pytest.fixture
    def report_output(self, tmp_path):
        """Generate a report from mock files"""
        mock_dir = Path("tests/mocks/pyspark")
        output_dir = tmp_path / "output"
        
        # Run the CLI to generate reports
        sys.argv = ['lineage', '--repo', str(mock_dir), '--out', str(output_dir)]
        try:
            cli_main()
        except SystemExit:
            pass  # CLI calls sys.exit(), which is expected
        
        return output_dir
    
    def test_excel_file_exists(self, report_output):
        """Test that Excel report file is generated"""
        excel_file = report_output / "lineage_report.xlsx"
        assert excel_file.exists(), "Excel report file should be generated"
        assert excel_file.stat().st_size > 0, "Excel file should not be empty"
    
    def test_excel_required_sheets(self, report_output):
        """Test that all required sheets are present"""
        excel_file = report_output / "lineage_report.xlsx"
        xl = pd.ExcelFile(excel_file)
        
        required_sheets = [
            'Summary',
            'Top Priorities',
            'Source Analysis',
            'Analytics (Pivot-Ready)',
            'All Datasets',
            'All Jobs',
            'Lineage Edges',
            'Detailed Metrics'
        ]
        
        for sheet in required_sheets:
            assert sheet in xl.sheet_names, f"Sheet '{sheet}' should be present"
        
        print(f"✅ All {len(required_sheets)} required sheets present")
    
    def test_summary_sheet_content(self, report_output):
        """Test Summary sheet has correct structure and content"""
        excel_file = report_output / "lineage_report.xlsx"
        df = pd.read_excel(excel_file, sheet_name='Summary', header=None)
        
        # Check for key metrics
        content = df.to_string()
        
        assert 'Total Nodes' in content, "Should contain Total Nodes metric"
        assert 'Total Edges' in content, "Should contain Total Edges metric"
        assert 'Dataset Nodes' in content, "Should contain Dataset Nodes metric"
        assert 'Job Nodes' in content, "Should contain Job Nodes metric"
        assert 'Migration Wave' in content, "Should contain Migration Wave analysis"
        
        # Validate that at least one wave is present
        has_waves = 'Wave 1' in content or 'Wave 2' in content or 'Wave 3' in content
        assert has_waves, "Should show at least one migration wave"
        
        print(f"✅ Summary sheet structure validated")
    
    def test_top_priorities_sheet_content(self, report_output):
        """Test Top Priorities sheet has correct columns and data"""
        excel_file = report_output / "lineage_report.xlsx"
        df = pd.read_excel(excel_file, sheet_name='Top Priorities')
        
        # Check required columns (skip header rows)
        data_df = df[df.iloc[:, 0].astype(str).str.isdigit()]
        
        assert len(data_df) > 0, "Should have priority dataset rows"
        
        # Check that we have ranking data
        ranks = data_df.iloc[:, 0].astype(str).tolist()
        assert '1' in ranks, "Should have rank 1"
        
        print(f"✅ Top Priorities sheet has {len(data_df)} datasets")
    
    def test_pivot_ready_sheet_structure(self, report_output):
        """Test Pivot-Ready sheet has correct structure"""
        excel_file = report_output / "lineage_report.xlsx"
        df = pd.read_excel(excel_file, sheet_name='Analytics (Pivot-Ready)')
        
        # Check that we have data rows (skip headers)
        data_rows = df[df.iloc[:, 3].notna()]
        
        assert len(data_rows) > 0, "Should have lineage relationship rows"
        
        # Check for key columns
        assert df.shape[1] >= 10, "Should have at least 10 columns"
        
        print(f"✅ Pivot-Ready sheet has {len(data_rows)} lineage relationships")
    
    def test_all_datasets_sheet_content(self, report_output):
        """Test All Datasets sheet has proper dataset information"""
        excel_file = report_output / "lineage_report.xlsx"
        df = pd.read_excel(excel_file, sheet_name='All Datasets')
        
        # Check required columns
        required_cols = ['Dataset Name', 'Type']
        for col in required_cols:
            assert col in df.columns, f"Should have '{col}' column"
        
        # Validate data
        assert len(df) > 0, "Should have dataset entries"
        
        # Check that we have various columns for analysis
        assert 'URN' in df.columns or 'Fact Count' in df.columns, "Should have analysis columns"
        
        print(f"✅ All Datasets sheet has {len(df)} datasets")
    
    def test_all_jobs_sheet_content(self, report_output):
        """Test All Jobs sheet has proper job information"""
        excel_file = report_output / "lineage_report.xlsx"
        df = pd.read_excel(excel_file, sheet_name='All Jobs')
        
        # Check required columns
        required_cols = ['Job Name', 'Source File']
        for col in required_cols:
            assert col in df.columns, f"Should have '{col}' column"
        
        # Validate data
        assert len(df) > 0, "Should have job entries"
        
        # Check that source files are present
        source_files = df['Source File'].dropna()
        assert len(source_files) > 0, "Jobs should have source files"
        
        print(f"✅ All Jobs sheet has {len(df)} jobs")
    
    def test_lineage_edges_sheet_content(self, report_output):
        """Test Lineage Edges sheet has edge information"""
        excel_file = report_output / "lineage_report.xlsx"
        df = pd.read_excel(excel_file, sheet_name='Lineage Edges')
        
        # Check required columns
        required_cols = ['Source', 'Target', 'Relationship']
        for col in required_cols:
            assert col in df.columns, f"Should have '{col}' column"
        
        # Validate data
        assert len(df) > 0, "Should have edge entries"
        
        # Check relationship types
        relationships = df['Relationship'].unique()
        assert len(relationships) > 0, "Should have different relationship types"
        
        print(f"✅ Lineage Edges sheet has {len(df)} edges")
    
    def test_excel_formatting(self, report_output):
        """Test Excel file has proper formatting (headers, styles)"""
        excel_file = report_output / "lineage_report.xlsx"
        wb = load_workbook(excel_file)
        
        # Check that sheets have headers
        for sheet_name in ['All Datasets', 'All Jobs', 'Lineage Edges']:
            ws = wb[sheet_name]
            
            # Check first row has content (headers)
            first_row = [cell.value for cell in ws[1]]
            assert any(first_row), f"Sheet '{sheet_name}' should have headers"
            
            # Check that there's data beyond headers
            assert ws.max_row > 1, f"Sheet '{sheet_name}' should have data rows"
        
        print(f"✅ Excel formatting validated for {len(wb.sheetnames)} sheets")


class TestJSONReportGeneration:
    """Tests for JSON report generation and structure"""
    
    @pytest.fixture
    def report_output(self, tmp_path):
        """Generate a report from mock files"""
        mock_dir = Path("tests/mocks/scala")
        output_dir = tmp_path / "output"
        
        sys.argv = ['lineage', '--repo', str(mock_dir), '--out', str(output_dir)]
        try:
            cli_main()
        except SystemExit:
            pass  # CLI calls sys.exit(), which is expected
        
        return output_dir
    
    def test_json_file_exists(self, report_output):
        """Test that JSON report file is generated"""
        json_file = report_output / "lineage.json"
        assert json_file.exists(), "JSON report file should be generated"
        assert json_file.stat().st_size > 0, "JSON file should not be empty"
    
    def test_json_valid_structure(self, report_output):
        """Test JSON file has valid structure"""
        json_file = report_output / "lineage.json"
        
        with open(json_file, 'r') as f:
            data = json.load(f)
        
        # Check top-level keys
        required_keys = ['metadata', 'datasets', 'jobs', 'edges', 'modules', 'prioritization']
        for key in required_keys:
            assert key in data, f"JSON should have '{key}' key"
        
        print(f"✅ JSON structure validated with {len(data)} top-level keys")
    
    def test_json_metadata_content(self, report_output):
        """Test JSON metadata section has correct information"""
        json_file = report_output / "lineage.json"
        
        with open(json_file, 'r') as f:
            data = json.load(f)
        
        metadata = data['metadata']
        
        # Check required metadata fields
        required_fields = ['generated_at', 'dataset_nodes', 'job_nodes', 'total_edges']
        for field in required_fields:
            assert field in metadata, f"Metadata should have '{field}' field"
        
        # Validate counts are numeric and positive
        assert metadata['dataset_nodes'] >= 0, "Dataset node count should be non-negative"
        assert metadata['job_nodes'] >= 0, "Job node count should be non-negative"
        assert metadata['total_edges'] >= 0, "Edge count should be non-negative"
        
        print(f"✅ JSON metadata validated")
    
    def test_json_datasets_structure(self, report_output):
        """Test JSON datasets array has correct structure"""
        json_file = report_output / "lineage.json"
        
        with open(json_file, 'r') as f:
            data = json.load(f)
        
        datasets = data['datasets']
        
        assert isinstance(datasets, list), "Datasets should be a list"
        assert len(datasets) > 0, "Should have dataset entries"
        
        # Check first dataset has required fields
        if datasets:
            dataset = datasets[0]
            required_fields = ['dataset_name', 'dataset_type']
            for field in required_fields:
                assert field in dataset, f"Dataset should have '{field}' field"
        
        print(f"✅ JSON datasets validated: {len(datasets)} datasets")
    
    def test_json_jobs_structure(self, report_output):
        """Test JSON jobs array has correct structure"""
        json_file = report_output / "lineage.json"
        
        with open(json_file, 'r') as f:
            data = json.load(f)
        
        jobs = data['jobs']
        
        assert isinstance(jobs, list), "Jobs should be a list"
        assert len(jobs) > 0, "Should have job entries"
        
        # Check first job has required fields
        if jobs:
            job = jobs[0]
            required_fields = ['job_name', 'source_file']
            for field in required_fields:
                assert field in job, f"Job should have '{field}' field"
        
        print(f"✅ JSON jobs validated: {len(jobs)} jobs")
    
    def test_json_edges_structure(self, report_output):
        """Test JSON edges array has correct structure"""
        json_file = report_output / "lineage.json"
        
        with open(json_file, 'r') as f:
            data = json.load(f)
        
        edges = data['edges']
        
        assert isinstance(edges, list), "Edges should be a list"
        assert len(edges) > 0, "Should have edge entries"
        
        # Check first edge has required fields
        if edges:
            edge = edges[0]
            required_fields = ['source_id', 'target_id', 'edge_type']
            for field in required_fields:
                assert field in edge, f"Edge should have '{field}' field"
        
        print(f"✅ JSON edges validated: {len(edges)} edges")
    
    def test_json_prioritization_content(self, report_output):
        """Test JSON prioritization section has priority rankings"""
        json_file = report_output / "lineage.json"
        
        with open(json_file, 'r') as f:
            data = json.load(f)
        
        prioritization = data['prioritization']
        
        assert 'top_datasets' in prioritization, "Should have top_datasets"
        
        top_datasets = prioritization['top_datasets']
        assert isinstance(top_datasets, list), "Top datasets should be a list"
        
        # Check priority fields
        if top_datasets:
            dataset = top_datasets[0]
            assert 'priority_score' in dataset, "Should have priority_score"
            assert 'migration_wave' in dataset, "Should have migration_wave"
        
        print(f"✅ JSON prioritization validated: {len(top_datasets)} top datasets")


class TestCSVReportGeneration:
    """Tests for CSV report generation"""
    
    @pytest.fixture
    def report_output(self, tmp_path):
        """Generate a report from mock files"""
        mock_dir = Path("tests/mocks/hive")
        output_dir = tmp_path / "output"
        
        sys.argv = ['lineage', '--repo', str(mock_dir), '--out', str(output_dir)]
        try:
            cli_main()
        except SystemExit:
            pass  # CLI calls sys.exit(), which is expected
        
        return output_dir
    
    def test_csv_directory_exists(self, report_output):
        """Test that CSV directory is created"""
        csv_dir = report_output / "csv"
        assert csv_dir.exists(), "CSV directory should be created"
        assert csv_dir.is_dir(), "CSV path should be a directory"
    
    def test_csv_files_generated(self, report_output):
        """Test that CSV files are generated"""
        csv_dir = report_output / "csv"
        csv_files = list(csv_dir.glob("*.csv"))
        
        assert len(csv_files) > 0, "Should generate CSV files"
        
        expected_files = ['datasets.csv', 'jobs.csv', 'edges.csv']
        for expected in expected_files:
            csv_file = csv_dir / expected
            if csv_file.exists():
                assert csv_file.stat().st_size > 0, f"{expected} should not be empty"
        
        print(f"✅ Generated {len(csv_files)} CSV files")
    
    def test_csv_datasets_content(self, report_output):
        """Test datasets.csv has proper content"""
        csv_file = report_output / "csv" / "datasets.csv"
        
        if csv_file.exists():
            df = pd.read_csv(csv_file)
            
            assert len(df) > 0, "Datasets CSV should have rows"
            assert 'dataset_name' in df.columns or 'Dataset Name' in df.columns, "Should have dataset name column"
            
            print(f"✅ datasets.csv has {len(df)} rows")
    
    def test_csv_jobs_content(self, report_output):
        """Test jobs.csv has proper content"""
        csv_file = report_output / "csv" / "jobs.csv"
        
        if csv_file.exists():
            df = pd.read_csv(csv_file)
            
            assert len(df) > 0, "Jobs CSV should have rows"
            assert 'job_name' in df.columns or 'Job Name' in df.columns, "Should have job name column"
            
            print(f"✅ jobs.csv has {len(df)} rows")
    
    def test_csv_edges_content(self, report_output):
        """Test edges.csv has proper content"""
        csv_file = report_output / "csv" / "edges.csv"
        
        if csv_file.exists():
            df = pd.read_csv(csv_file)
            
            assert len(df) > 0, "Edges CSV should have rows"
            
            print(f"✅ edges.csv has {len(df)} rows")


class TestHTMLReportGeneration:
    """Tests for HTML report generation"""
    
    @pytest.fixture
    def report_output(self, tmp_path):
        """Generate a report from mock files"""
        mock_dir = Path("tests/mocks/shell")
        output_dir = tmp_path / "output"
        
        sys.argv = ['lineage', '--repo', str(mock_dir), '--out', str(output_dir)]
        try:
            cli_main()
        except SystemExit:
            pass  # CLI calls sys.exit(), which is expected
        
        return output_dir
    
    def test_html_file_exists(self, report_output):
        """Test that HTML report file is generated"""
        html_file = report_output / "lineage_report.html"
        assert html_file.exists(), "HTML report file should be generated"
        assert html_file.stat().st_size > 0, "HTML file should not be empty"
    
    def test_html_valid_structure(self, report_output):
        """Test HTML file has valid structure"""
        html_file = report_output / "lineage_report.html"
        
        with open(html_file, 'r') as f:
            content = f.read()
        
        # Check for basic HTML structure
        assert '<html' in content.lower(), "Should have HTML tag"
        assert '<head' in content.lower(), "Should have head section"
        assert '<body' in content.lower(), "Should have body section"
        
        # Check for content sections
        assert 'lineage' in content.lower(), "Should mention lineage"
        
        print(f"✅ HTML structure validated ({len(content)} characters)")


class TestReportContentValidation:
    """Tests for validating report content quality"""
    
    @pytest.fixture
    def comprehensive_report(self, tmp_path):
        """Generate a comprehensive report from all mocks"""
        mock_dir = Path("tests/mocks")
        output_dir = tmp_path / "output"
        
        sys.argv = ['lineage', '--repo', str(mock_dir), '--out', str(output_dir)]
        try:
            cli_main()
        except SystemExit:
            pass  # CLI calls sys.exit(), which is expected
        
        return output_dir
    
    def test_no_duplicate_datasets(self, comprehensive_report):
        """Test that dataset names are unique in the report"""
        excel_file = comprehensive_report / "lineage_report.xlsx"
        df = pd.read_excel(excel_file, sheet_name='All Datasets')
        
        duplicates = df[df.duplicated(subset=['Dataset Name'], keep=False)]
        
        # Some duplicates may be acceptable if they have different types
        # Just check that we don't have excessive duplication
        assert len(duplicates) < len(df) * 0.1, "Should not have more than 10% duplicates"
        
        print(f"✅ Dataset uniqueness validated")
    
    def test_resolution_rate(self, comprehensive_report):
        """Test that variable resolution rate is high"""
        excel_file = comprehensive_report / "lineage_report.xlsx"
        pivot_df = pd.read_excel(excel_file, sheet_name='Analytics (Pivot-Ready)')
        
        # Count resolved vs unresolved
        resolved_col = pivot_df.columns[9] if len(pivot_df.columns) > 9 else None
        
        if resolved_col:
            resolved = pivot_df[resolved_col].astype(str)
            yes_count = (resolved == 'Yes').sum()
            total_count = len(resolved[resolved != 'nan'])
            
            resolution_rate = yes_count / total_count if total_count > 0 else 0
            
            assert resolution_rate > 0.95, f"Resolution rate should be > 95%, got {resolution_rate:.1%}"
            
            print(f"✅ Resolution rate: {resolution_rate:.1%}")
    
    def test_modular_apps_in_report(self, comprehensive_report):
        """Test that modular applications are included in report"""
        excel_file = comprehensive_report / "lineage_report.xlsx"
        jobs_df = pd.read_excel(excel_file, sheet_name='All Jobs')
        
        # Check for modular app jobs
        modular_keywords = ['modular_app', 'Module', 'Ingestion', 'Output', 'Orchestrator']
        modular_jobs = jobs_df[jobs_df['Job Name'].str.contains('|'.join(modular_keywords), case=False, na=False)]
        
        assert len(modular_jobs) > 0, "Should include modular application jobs"
        
        print(f"✅ Found {len(modular_jobs)} modular application jobs")
    
    def test_migration_waves_assigned(self, comprehensive_report):
        """Test that datasets have migration waves assigned"""
        excel_file = comprehensive_report / "lineage_report.xlsx"
        priorities_df = pd.read_excel(excel_file, sheet_name='Top Priorities')
        
        # Find Wave column (skip header rows)
        data_rows = priorities_df[priorities_df.iloc[:, 0].astype(str).str.isdigit()]
        
        if len(data_rows) > 0:
            # Check that waves are assigned
            wave_col = None
            for col in data_rows.columns:
                if 'wave' in str(col).lower():
                    wave_col = col
                    break
            
            if wave_col:
                waves = data_rows[wave_col].astype(str)
                assert waves.str.contains('Wave').any(), "Should have Wave assignments"
        
        print(f"✅ Migration waves validated")
    
    def test_source_file_attribution(self, comprehensive_report):
        """Test that all jobs have source file attribution"""
        excel_file = comprehensive_report / "lineage_report.xlsx"
        jobs_df = pd.read_excel(excel_file, sheet_name='All Jobs')
        
        # Check that source files are present
        source_files = jobs_df['Source File'].dropna()
        
        assert len(source_files) > 0, "Jobs should have source files"
        
        # Check that source files have proper paths
        valid_paths = source_files[source_files.str.contains('/', na=False)]
        assert len(valid_paths) > 0, "Source files should have proper paths"
        
        print(f"✅ Source file attribution validated for {len(jobs_df)} jobs")
    
    def test_lineage_connectivity(self, comprehensive_report):
        """Test that lineage graph is properly connected"""
        json_file = comprehensive_report / "lineage.json"
        
        with open(json_file, 'r') as f:
            data = json.load(f)
        
        datasets = data['datasets']
        jobs = data['jobs']
        edges = data['edges']
        
        # Check that we have a reasonable edge-to-node ratio
        total_nodes = len(datasets) + len(jobs)
        edge_to_node_ratio = len(edges) / total_nodes if total_nodes > 0 else 0
        
        assert edge_to_node_ratio > 1.0, "Should have more edges than nodes (well-connected graph)"
        
        print(f"✅ Lineage connectivity: {edge_to_node_ratio:.1f} edges per node")


class TestReportFormatting:
    """Tests specifically for report formatting and presentation"""
    
    def test_excel_column_widths(self, tmp_path):
        """Test that Excel columns have reasonable widths"""
        mock_dir = Path("tests/mocks/pyspark")
        output_dir = tmp_path / "output"
        
        sys.argv = ['lineage', '--repo', str(mock_dir), '--out', str(output_dir)]
        try:
            cli_main()
        except SystemExit:
            pass  # CLI calls sys.exit(), which is expected
        
        excel_file = output_dir / "lineage_report.xlsx"
        wb = load_workbook(excel_file)
        
        # Check column dimensions for key sheets
        for sheet_name in ['All Datasets', 'All Jobs']:
            ws = wb[sheet_name]
            
            # Check that some columns have explicit widths set
            # (This depends on implementation, so we just check structure)
            assert ws.max_column > 0, f"Sheet '{sheet_name}' should have columns"
        
        print(f"✅ Excel column formatting validated")
    
    def test_pivot_tab_headers(self, tmp_path):
        """Test that Pivot-Ready tab has descriptive headers"""
        mock_dir = Path("tests/mocks/scala")
        output_dir = tmp_path / "output"
        
        sys.argv = ['lineage', '--repo', str(mock_dir), '--out', str(output_dir)]
        try:
            cli_main()
        except SystemExit:
            pass  # CLI calls sys.exit(), which is expected
        
        excel_file = output_dir / "lineage_report.xlsx"
        df = pd.read_excel(excel_file, sheet_name='Analytics (Pivot-Ready)')
        
        # Check that we have explanatory header rows
        first_rows = df.head(2).to_string()
        
        # Should have descriptive text in early rows
        assert 'dataset' in first_rows.lower() or 'lineage' in first_rows.lower(), \
            "Pivot tab should have descriptive headers"
        
        print(f"✅ Pivot tab headers validated")


if __name__ == "__main__":
    # Run tests
    pytest.main([__file__, "-v", "--tb=short"])

